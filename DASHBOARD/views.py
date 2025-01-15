from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.views.generic.edit import UpdateView,CreateView,DeleteView
from .models import Company,Employee,Convention,Setting,Folder
from .forms import CompanyForm,EmployeeForm,ConventionForm,SettingForm,FolderForm
from django.http import HttpResponse
# Pour l'exportation des donnees au format excel
import openpyxl
# Pour l'exportation des donnees au format pdf
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


def verify_string(chaine):
    char_in=False
    for char in chaine:
        if char in ['<','>','@','/','$','?','!','*']:
            char_in=True
            break
    return char_in

@login_required # Force l'utilisateur a se loguer pour acceder a son dashboard
def dashboard(request):
    user_employees = Employee.objects.filter(added_by_user=request.user)
    user_companies = Company.objects.filter(added_by_user=request.user)
    user_conventions = Convention.objects.filter(added_by_user=request.user)
    user_settings = Setting.objects.filter(added_by_user=request.user)
    user_folders = Folder.objects.filter(added_by_user=request.user)
    return render(request, 'dashboard.html',
                  {
                    'employees':user_employees,'companies':user_companies,'conventions':user_conventions,
                    'settings':user_settings, 'folders':user_folders
                  })

@login_required # Force l'utilisateur a se loguer pour acceder a la vue des recherches de dossiers
def search(request):
    saisie=request.GET['folder']
    if saisie != "" and len(saisie)<=3 and verify_string(saisie)==False:
        return render(request, 'search.html',
                    {'folders':Folder.objects.filter(added_by_user=request.user,status__icontains=saisie)})
    else:
        return redirect('dashboard')

# Fonction d'exportation des donnees au format excel/pdf
def export_folder_data(request,folder_id):
    # Recuperation du dossier conserné
    folder = Folder.objects.get(id=folder_id)
    if request.GET['format'] == 'excel':
        # Creation du fichier excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Nom de l'employé"
        ws['B1'] = folder.employee.name
        ws['A2'] = "Société d'appartenance"
        ws['B2'] = folder.employee.company.company_name
        ws['A3'] = "Secteur d'activité"
        ws['B3'] = folder.employee.company.get_activity_sector_display()
        ws['A4'] = "Convention de calcul de l'IFC"
        ws['B4'] = folder.setting.convention.get_name_display()
        ws['A5'] = "Statut du dossier"
        ws['B5'] = folder.get_status_display()
        ws['A6'] = "Resultat de l'IFC"
        ws['B6'] = folder.ifc_result
        # Creation de la reponse HTTP pour telecharger le fichier excel
        response = HttpResponse(content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Rapport {folder.employee.name}.xlsx"'
        # Sauvegarde du fichier
        wb.save(response)
    elif request.GET['format'] == 'pdf':
        # Creation de la reponse HTTP pour telecharger le fichier pdf
        response = HttpResponse(content_type = 'application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Rapport {folder.employee.name}.pdf"'
        # Creation du fichier pdf
        c = canvas.Canvas(response,pagesize=letter)
        c.drawString(100,750,f"Rapport pour l'employé: {folder.employee.name}")
        c.drawString(100,730,f"De la société: {folder.employee.company.company_name}")
        c.drawString(100,710,f"Qui fait dans: {folder.employee.company.get_activity_sector_display()}")
        c.drawString(100,690,f"Statut du dossier: {folder.get_status_display()}")
        c.drawString(100,670,f"Convention de calcul de l'IFC: {folder.setting.convention.get_name_display()}")
        c.drawString(100,650,f"Valeur calculée de l'IFC: {folder.ifc_result} FCFA")

        # Finalisation du PDF
        c.showPage()
        c.save()
    
    return response


class AddCompany(LoginRequiredMixin,CreateView):
    model = Company
    form_class = CompanyForm
    template_name = 'add_element.html'
    success_url = '/dashboard'

    def form_valid(self, form):
        form.instance.added_by_user=self.request.user
        return super().form_valid(form)
    
class UpdateCompany(LoginRequiredMixin,UpdateView):
    model = Company
    form_class = CompanyForm
    template_name = 'update_element.html'
    success_url = '/dashboard'

class DeleteCompany(LoginRequiredMixin,DeleteView):
    model = Company
    template_name = 'delete_element.html'
    success_url = '/dashboard'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.has_perm('DASHBOARD.delete_company'):
            print(f'Attention, l\'utilisateur {request.user} a tenté une suppression via la barre d\'adresse')
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)


class AddEmployee(LoginRequiredMixin,CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'add_element.html'
    success_url = '/dashboard'

    def form_valid(self, form):
        form.instance.added_by_user=self.request.user
        return super().form_valid(form)
    
class UpdateEmployee(LoginRequiredMixin,UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'update_element.html'
    success_url = '/dashboard'

class DeleteEmployee(LoginRequiredMixin,DeleteView):
    model = Employee
    template_name = 'delete_element.html'
    success_url = '/dashboard'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.has_perm('DASHBOARD.delete_company'):
            print(f'Attention, l\'utilisateur {request.user} a tenté une suppression via la barre d\'adresse')
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)


class AddConvention(LoginRequiredMixin,CreateView):
    model = Convention
    form_class = ConventionForm
    template_name = 'add_element.html'
    success_url = '/dashboard'

    def form_valid(self, form):
        form.instance.added_by_user=self.request.user
        return super().form_valid(form)
    
class UpdateConvention(LoginRequiredMixin,UpdateView):
    model = Convention
    form_class = ConventionForm
    template_name = 'update_element.html'
    success_url = '/dashboard'

class DeleteConvention(LoginRequiredMixin,DeleteView):
    model = Convention
    template_name = 'delete_element.html'
    success_url = '/dashboard'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.has_perm('DASHBOARD.delete_company'):
            print(f'Attention, l\'utilisateur {request.user} a tenté une suppression via la barre d\'adresse')
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)


class AddSetting(LoginRequiredMixin,CreateView):
    model = Setting
    form_class = SettingForm
    template_name = 'add_element.html'
    success_url = '/dashboard'

    def form_valid(self, form):
        form.instance.added_by_user=self.request.user
        return super().form_valid(form)
    
class UpdateSetting(LoginRequiredMixin,UpdateView):
    model = Setting
    form_class = SettingForm
    template_name = 'update_element.html'
    success_url = '/dashboard'

class DeleteSetting(LoginRequiredMixin,DeleteView):
    model = Setting
    template_name = 'delete_element.html'
    success_url = '/dashboard'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.has_perm('DASHBOARD.delete_company'):
            print(f'Attention, l\'utilisateur {request.user} a tenté une suppression via la barre d\'adresse')
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)


class AddFolder(LoginRequiredMixin,CreateView):
    model = Folder
    form_class = FolderForm
    template_name = 'add_element.html'
    success_url = '/dashboard'

    def form_valid(self, form):
        form.instance.added_by_user=self.request.user
        # Calcul de l'IFC avant sauvegarde du dossier dans la BD
        folder = form.save(commit=False)  # Ne pas sauvegarder tout de suite
        folder.employee_ifc_compute()  # Appel la méthode de calcul de l'IFC
        folder.save()  # Sauvegarde de l'objet avec le résultat de l'IFC
        return super().form_valid(form)
    
class UpdateFolder(LoginRequiredMixin,UpdateView):
    model = Folder
    form_class = FolderForm
    template_name = 'update_element.html'
    success_url = '/dashboard'

    def form_valid(self, form):
        # On recalcule l'IFC avant de sauvegarder les modifications du dossier dans la BD
        folder = form.save(commit=False)  # Ne pas sauvegarder tout de suite
        folder.employee_ifc_compute()  # Appel la méthode de calcul de l'IFC
        folder.save()  # Sauvegarde de l'objet avec le nouveau résultat de l'IFC
        return super().form_valid(form)

class DeleteFolder(LoginRequiredMixin,DeleteView):
    model = Folder
    template_name = 'delete_element.html'
    success_url = '/dashboard'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.has_perm('DASHBOARD.delete_company'):
            print(f'Attention, l\'utilisateur {request.user} a tenté une suppression via la barre d\'adresse')
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)